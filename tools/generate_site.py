from __future__ import annotations

import csv
import html
import json
import re
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from urllib.parse import urlparse

from openpyxl import load_workbook
from PIL import Image, ImageDraw, ImageFont


ROOT = Path(__file__).resolve().parents[1]
SOURCE = ROOT / "ai-api-sites-table.xlsx"
SITE_URL = "https://bubblevv.github.io/ai-api-gongyi-nav/"
OUTPUTS = {
    "index": ROOT / "index.html",
    "table_html": ROOT / "ai-api-sites-table.html",
    "share_html": ROOT / "ai-api-sites-share.html",
    "readme": ROOT / "README.md",
    "markdown": ROOT / "ai-api-sites-table.md",
    "csv": ROOT / "ai-api-sites-table.csv",
    "json": ROOT / "data" / "sites.json",
    "cover": ROOT / "assets" / "ai-api-gongyi-nav-cover.png",
    "robots": ROOT / "robots.txt",
    "sitemap": ROOT / "sitemap.xml",
}

KEYWORDS = [
    "ai api",
    "AI API",
    "AI API 公益站",
    "公益中转站",
    "公益 API",
    "AI API 中转站",
    "AI 中转站",
    "API 中转站",
    "GPT",
    "GPT API",
    "GPT API 免费",
    "OpenAI API",
    "OpenAI API 中转",
    "Claude",
    "Claude API",
    "Claude API 中转",
    "Codex",
    "Codex 编程",
    "DeepSeek",
    "DeepSeek API",
    "DeepSeek API 免费",
    "Gemini",
    "Gemini API",
    "GLM",
    "GLM API",
    "MiniMax",
    "New API",
    "New API 公益站",
    "ChatGPT API",
    "免费 AI API",
    "AI API 导航",
    "免费 API 额度",
    "API 签到额度",
    "编程",
    "编程 API",
    "龙虾",
    "酒馆",
    "酒馆 API",
]

FILTER_TAGS = ["公益", "Claude", "GPT", "DeepSeek", "生图", "稳定", "注册赠送", "签到"]
AFF_SIGNALS = ("aff=", "/invite/", "invite=")


@dataclass(frozen=True)
class Site:
    index: int
    name: str
    url: str
    tags: tuple[str, ...]
    note: str
    added_date: str = ""

    @property
    def domain(self) -> str:
        host = urlparse(self.url).netloc
        return host.removeprefix("www.")

    @property
    def search_text(self) -> str:
        return " ".join([self.name, self.domain, self.note, self.added_date, *self.tags]).lower()

    @property
    def caution(self) -> bool:
        caution_words = ("不稳定", "不可用", "无法", "取消", "难抢", "需手机", "需加群")
        return any(word in self.note or any(word in tag for tag in self.tags) for word in caution_words)


def text(value: object) -> str:
    return "" if value is None else str(value).strip()


def split_tags(raw: str) -> tuple[str, ...]:
    return tuple(part.strip() for part in re.split(r"[;；]", raw) if part.strip())


def source_updated() -> str:
    mtime = datetime.fromtimestamp(SOURCE.stat().st_mtime)
    return mtime.strftime("%Y-%m-%d")


def load_sites() -> tuple[str, str, list[Site]]:
    workbook = load_workbook(SOURCE, data_only=True)
    sheet = workbook.active
    heading = text(sheet.cell(row=1, column=1).value)
    heading_lines = [line.strip() for line in heading.splitlines() if line.strip()]
    title = heading_lines[0] if heading_lines else "AI API 公益中转站导航"
    subtitle = " ".join(heading_lines[1:])
    sites: list[Site] = []

    header_row = None
    for row_number in range(1, min(sheet.max_row, 10) + 1):
        headers = [text(sheet.cell(row=row_number, column=column).value) for column in range(1, 6)]
        if headers[:5] == ["序号", "站点", "链接", "标签", "备注"]:
            header_row = row_number
            break
    if header_row is None:
        raise ValueError("Could not find header row: 序号/站点/链接/标签/备注")

    for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
        if not any(row):
            continue
        index_raw, name_raw, url_raw, tags_raw, note_raw = row[:5]
        added_date_raw = row[5] if len(row) > 5 else ""
        name = text(name_raw)
        url = text(url_raw)
        if not name or not url:
            continue
        parsed = urlparse(url)
        if parsed.scheme not in {"http", "https"} or not parsed.netloc:
            raise ValueError(f"Invalid URL for {name}: {url}")
        if not any(signal in url for signal in AFF_SIGNALS):
            raise ValueError(f"Referral signal missing for {name}: {url}")
        sites.append(
            Site(
                index=int(index_raw),
                name=name,
                url=url,
                tags=split_tags(text(tags_raw)),
                note=text(note_raw),
                added_date=text(added_date_raw),
            )
        )

    sites.sort(key=lambda site: site.index)
    return title, subtitle, sites


def esc(value: object) -> str:
    return html.escape(str(value), quote=True)


def write_cover(title: str, sites: list[Site]) -> None:
    OUTPUTS["cover"].parent.mkdir(parents=True, exist_ok=True)
    image = Image.new("RGB", (1200, 630), "#f7faf7")
    draw = ImageDraw.Draw(image)

    font_dir = Path("C:/Windows/Fonts")
    font_regular = font_dir / "msyh.ttc"
    font_bold = font_dir / "msyhbd.ttc"
    title_font = ImageFont.truetype(str(font_bold if font_bold.exists() else font_regular), 58)
    sub_font = ImageFont.truetype(str(font_regular), 28)
    card_font = ImageFont.truetype(str(font_regular), 23)
    small_font = ImageFont.truetype(str(font_regular), 20)

    draw.rectangle((0, 0, 1200, 630), fill="#f7faf7")
    draw.rectangle((0, 0, 1200, 18), fill="#12805c")
    draw.rectangle((0, 18, 1200, 32), fill="#2457a6")
    draw.rectangle((0, 32, 1200, 42), fill="#d98b2b")
    draw.text((64, 78), title, fill="#15231f", font=title_font)
    draw.text((68, 156), "GPT / Claude / Codex / DeepSeek / 免费额度 / API 中转", fill="#38524a", font=sub_font)
    draw.text((68, 207), f"收录 {len(sites)} 个 AI API 公益站和中转站入口", fill="#66736f", font=small_font)

    columns = [(68, 284), (438, 284), (808, 284)]
    for idx, site in enumerate(sites[:9]):
        col_x, base_y = columns[idx % 3]
        y = base_y + (idx // 3) * 92
        draw.rounded_rectangle((col_x, y, col_x + 320, y + 64), radius=8, fill="#ffffff", outline="#d9e4de", width=2)
        draw.rectangle((col_x, y, col_x + 7, y + 64), fill=("#12805c" if not site.caution else "#d98b2b"))
        draw.text((col_x + 20, y + 10), site.name[:15], fill="#182821", font=card_font)
        draw.text((col_x + 20, y + 38), site.domain[:26], fill="#66736f", font=small_font)

    image.save(OUTPUTS["cover"], optimize=True)


def write_csv(sites: list[Site]) -> None:
    with OUTPUTS["csv"].open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(["序号", "站点", "AFF链接", "标签", "备注", "添加日期"])
        for site in sites:
            writer.writerow([site.index, site.name, site.url, ";".join(site.tags), site.note, site.added_date])


def site_to_dict(site: Site) -> dict[str, object]:
    return {
        "index": site.index,
        "name": site.name,
        "url": site.url,
        "domain": site.domain,
        "tags": list(site.tags),
        "note": site.note,
        "added_date": site.added_date,
    }


def write_json(title: str, subtitle: str, updated: str, sites: list[Site]) -> None:
    OUTPUTS["json"].parent.mkdir(parents=True, exist_ok=True)
    payload = {
        "title": title,
        "description": subtitle,
        "updated": updated,
        "keywords": KEYWORDS,
        "sites": [site_to_dict(site) for site in sites],
    }
    OUTPUTS["json"].write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def markdown_link(site: Site) -> str:
    return f"[{site.name}]({site.url})"


def write_markdown(title: str, subtitle: str, updated: str, sites: list[Site]) -> str:
    rows = [
        f"# {title}",
        "",
        f"> {subtitle}",
        "",
        f"更新日期：{updated}",
        "",
        "这里整理可用于编程、酒馆、龙虾和日常模型调用的 AI API 公益站与中转站入口，重点覆盖 GPT、Claude、DeepSeek、Gemini、GLM、MiniMax、Codex 等常见模型。",
        "",
        "使用前请自行确认额度、倍率、模型可用性和注册状态；不要上传隐私数据、密钥或敏感文件。本站只做导航收录，不对任何第三方站点的服务质量、稳定性或数据安全负责。",
        "",
        "| 序号 | 站点 | 专属链接 | 标签 | 备注 | 添加日期 |",
        "| --- | --- | --- | --- | --- | --- |",
    ]
    for site in sites:
        rows.append(
            f"| {site.index} | {site.name} | {markdown_link(site)} | {'; '.join(site.tags)} | {site.note} | {site.added_date} |"
        )
    content = "\n".join(rows) + "\n"
    OUTPUTS["readme"].write_text(content, encoding="utf-8")
    OUTPUTS["markdown"].write_text(content, encoding="utf-8")
    return content


def tag_html(tag: str) -> str:
    safe = esc(tag)
    return f'<span class="tag">{safe}</span>'


def card_html(site: Site) -> str:
    tags = "".join(tag_html(tag) for tag in site.tags)
    caution = " caution" if site.caution else ""
    added_date = (
        f'\n          <p class="added-date"><span>添加日期</span>{esc(site.added_date)}</p>'
        if site.added_date
        else ""
    )
    return f"""
        <article class="site-card{caution}" data-tags="{esc(' '.join(site.tags))}" data-search="{esc(site.search_text)}">
          <div class="site-main">
            <div>
              <p class="site-index">#{site.index}</p>
              <h2>{esc(site.name)}</h2>
              <p class="domain">{esc(site.domain)}</p>
            </div>
            <a class="visit" href="{esc(site.url)}" target="_blank" rel="sponsored noopener noreferrer" aria-label="访问 {esc(site.name)} 专属链接">访问</a>
          </div>
          <div class="tags">{tags}</div>{added_date}
          <p class="note">{esc(site.note)}</p>
        </article>"""


def write_html(title: str, subtitle: str, updated: str, sites: list[Site]) -> str:
    keyword_text = ", ".join(KEYWORDS)
    description = (
        "AI API 公益站和中转站导航，整理 GPT、Claude、DeepSeek、Gemini、"
        "Codex、New API 等接口入口，适合编程、酒馆、龙虾和免费额度检索。"
    )
    filters = "\n".join(
        f'<button class="filter" type="button" data-filter="{esc(tag)}">{esc(tag)}</button>' for tag in FILTER_TAGS
    )
    cards = "\n".join(card_html(site) for site in sites)
    json_ld = {
        "@context": "https://schema.org",
        "@type": "ItemList",
        "name": title,
        "description": description,
        "url": SITE_URL,
        "keywords": keyword_text,
        "dateModified": updated,
        "itemListElement": [
            {
                "@type": "ListItem",
                "position": position,
                "name": site.name,
                "url": site.url,
            }
            for position, site in enumerate(sites, start=1)
        ],
    }
    html_doc = f"""<!doctype html>
<html lang="zh-CN">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>{esc(title)} | AI API 公益站 GPT Claude DeepSeek 免费中转站</title>
  <meta name="description" content="{esc(description)}">
  <meta name="keywords" content="{esc(keyword_text)}">
  <link rel="canonical" href="{esc(SITE_URL)}">
  <link rel="icon" href="assets/ai-api-gongyi-nav-cover.png">
  <meta property="og:title" content="{esc(title)}">
  <meta property="og:description" content="{esc(description)}">
  <meta property="og:type" content="website">
  <meta property="og:image" content="assets/ai-api-gongyi-nav-cover.png">
  <meta property="og:url" content="{esc(SITE_URL)}">
  <script type="application/ld+json">{json.dumps(json_ld, ensure_ascii=False)}</script>
  <style>
    :root {{
      color-scheme: light;
      --ink: #16231f;
      --muted: #617069;
      --line: #dce5df;
      --paper: #fbfcfb;
      --panel: #ffffff;
      --green: #12805c;
      --blue: #2457a6;
      --amber: #c97920;
      --red: #b0443e;
      --shadow: 0 18px 50px rgba(35, 48, 42, 0.10);
    }}
    * {{ box-sizing: border-box; }}
    body {{
      margin: 0;
      font-family: "Microsoft YaHei", "PingFang SC", "Noto Sans CJK SC", system-ui, -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif;
      color: var(--ink);
      background:
        linear-gradient(180deg, rgba(18, 128, 92, 0.08), transparent 360px),
        var(--paper);
      letter-spacing: 0;
    }}
    a {{ color: inherit; }}
    .top-strip {{
      display: grid;
      grid-template-columns: 2fr 1.2fr 0.8fr;
      height: 12px;
    }}
    .top-strip span:nth-child(1) {{ background: var(--green); }}
    .top-strip span:nth-child(2) {{ background: var(--blue); }}
    .top-strip span:nth-child(3) {{ background: var(--amber); }}
    .wrap {{
      width: min(1160px, calc(100% - 32px));
      margin: 0 auto;
    }}
    header {{
      padding: 34px 0 22px;
    }}
    .hero {{
      display: grid;
      grid-template-columns: minmax(0, 1fr) minmax(280px, 420px);
      gap: 28px;
      align-items: center;
    }}
    .eyebrow {{
      margin: 0 0 10px;
      color: var(--green);
      font-size: 14px;
      font-weight: 700;
    }}
    h1 {{
      margin: 0;
      max-width: 820px;
      font-size: clamp(34px, 6vw, 64px);
      line-height: 1.05;
      letter-spacing: 0;
    }}
    .subtitle {{
      max-width: 780px;
      margin: 16px 0 0;
      color: var(--muted);
      font-size: 17px;
      line-height: 1.75;
    }}
    .cover {{
      display: block;
      width: 100%;
      height: auto;
      aspect-ratio: 1200 / 630;
      object-fit: contain;
      border: 1px solid var(--line);
      border-radius: 8px;
      box-shadow: var(--shadow);
      background: #f7faf7;
    }}
    .stats {{
      display: flex;
      flex-wrap: wrap;
      gap: 10px;
      margin-top: 18px;
    }}
    .stat {{
      min-width: 112px;
      padding: 10px 12px;
      border: 1px solid var(--line);
      border-radius: 8px;
      background: rgba(255, 255, 255, 0.82);
    }}
    .stat b {{
      display: block;
      font-size: 22px;
      line-height: 1.1;
    }}
    .stat span {{
      display: block;
      margin-top: 4px;
      color: var(--muted);
      font-size: 12px;
    }}
    .toolbar-band {{
      position: sticky;
      top: 0;
      z-index: 5;
      border-block: 1px solid var(--line);
      background: rgba(251, 252, 251, 0.94);
      backdrop-filter: blur(14px);
    }}
    .toolbar {{
      display: grid;
      grid-template-columns: minmax(220px, 1fr) auto;
      gap: 14px;
      align-items: center;
      padding: 14px 0;
    }}
    .search {{
      width: 100%;
      min-height: 44px;
      padding: 0 14px;
      border: 1px solid var(--line);
      border-radius: 8px;
      color: var(--ink);
      background: var(--panel);
      font: inherit;
      outline: none;
    }}
    .search:focus {{
      border-color: var(--blue);
      box-shadow: 0 0 0 3px rgba(36, 87, 166, 0.14);
    }}
    .filters {{
      display: flex;
      flex-wrap: wrap;
      gap: 8px;
      justify-content: flex-end;
    }}
    .filter {{
      min-height: 36px;
      padding: 0 11px;
      border: 1px solid var(--line);
      border-radius: 8px;
      color: var(--ink);
      background: var(--panel);
      font: inherit;
      cursor: pointer;
    }}
    .filter.active {{
      border-color: var(--green);
      color: #ffffff;
      background: var(--green);
    }}
    main {{
      padding: 28px 0 54px;
    }}
    .meta-row {{
      display: flex;
      justify-content: space-between;
      gap: 16px;
      margin-bottom: 16px;
      color: var(--muted);
      font-size: 14px;
    }}
    .grid {{
      display: grid;
      grid-template-columns: repeat(3, minmax(0, 1fr));
      gap: 14px;
    }}
    .site-card {{
      min-height: 220px;
      display: flex;
      flex-direction: column;
      justify-content: space-between;
      padding: 18px;
      border: 1px solid var(--line);
      border-top: 4px solid var(--green);
      border-radius: 8px;
      background: var(--panel);
      box-shadow: 0 10px 30px rgba(35, 48, 42, 0.06);
    }}
    .site-card.caution {{ border-top-color: var(--amber); }}
    .site-main {{
      display: grid;
      grid-template-columns: minmax(0, 1fr) auto;
      gap: 12px;
      align-items: start;
    }}
    .site-index {{
      margin: 0 0 6px;
      color: var(--muted);
      font-size: 13px;
    }}
    h2 {{
      margin: 0;
      overflow-wrap: anywhere;
      font-size: 22px;
      line-height: 1.25;
      letter-spacing: 0;
    }}
    .domain {{
      margin: 7px 0 0;
      color: var(--muted);
      overflow-wrap: anywhere;
      font-size: 13px;
    }}
    .visit {{
      display: inline-flex;
      align-items: center;
      justify-content: center;
      width: 58px;
      min-height: 38px;
      border-radius: 8px;
      color: #ffffff;
      background: var(--blue);
      text-decoration: none;
      font-size: 14px;
      font-weight: 700;
      white-space: nowrap;
    }}
    .visit:hover {{ background: #173f7d; }}
    .tags {{
      display: flex;
      flex-wrap: wrap;
      gap: 7px;
      margin: 16px 0 12px;
    }}
    .tag {{
      padding: 5px 7px;
      border: 1px solid #d8e2dc;
      border-radius: 6px;
      color: #315047;
      background: #f4f8f6;
      font-size: 12px;
      line-height: 1;
      overflow-wrap: anywhere;
    }}
    .added-date {{
      display: flex;
      align-items: center;
      gap: 7px;
      margin: -2px 0 10px;
      color: var(--muted);
      font-size: 13px;
      line-height: 1.4;
    }}
    .added-date span {{
      color: #315047;
      font-weight: 700;
    }}
    .note {{
      margin: 0;
      color: #3f4c47;
      font-size: 14px;
      line-height: 1.65;
    }}
    .empty {{
      display: none;
      padding: 36px 0;
      color: var(--muted);
      text-align: center;
    }}
    footer {{
      border-top: 1px solid var(--line);
      padding: 22px 0 34px;
      color: var(--muted);
      font-size: 13px;
      line-height: 1.7;
    }}
    @media (max-width: 920px) {{
      .hero {{ grid-template-columns: 1fr; }}
      .cover {{ max-width: 560px; }}
      .toolbar {{ grid-template-columns: 1fr; }}
      .filters {{ justify-content: flex-start; }}
      .grid {{ grid-template-columns: repeat(2, minmax(0, 1fr)); }}
    }}
    @media (max-width: 620px) {{
      .wrap {{ width: min(100% - 22px, 1160px); }}
      header {{ padding-top: 24px; }}
      h1 {{ font-size: 36px; }}
      .subtitle {{ font-size: 15px; }}
      .meta-row {{ display: block; }}
      .grid {{ grid-template-columns: 1fr; }}
      .site-card {{ min-height: 0; }}
      .filter {{ flex: 1 1 calc(50% - 8px); }}
    }}
  </style>
</head>
<body>
  <div class="top-strip" aria-hidden="true"><span></span><span></span><span></span></div>
  <header class="wrap">
    <div class="hero">
      <div>
        <p class="eyebrow">AI API / GPT / Claude / Codex / DeepSeek</p>
        <h1>{esc(title)}</h1>
        <p class="subtitle">{esc(subtitle)} 支持按模型、额度、倍率、签到、酒馆、编程等标签快速筛选。</p>
        <div class="stats" aria-label="站点概览">
          <div class="stat"><b>{len(sites)}</b><span>收录站点</span></div>
          <div class="stat"><b>{sum('公益' in site.tags for site in sites)}</b><span>公益标签</span></div>
          <div class="stat"><b>{sum('Claude' in site.tags for site in sites)}</b><span>Claude 标签</span></div>
          <div class="stat"><b>{updated}</b><span>更新日期</span></div>
        </div>
      </div>
      <img class="cover" src="assets/ai-api-gongyi-nav-cover.png" alt="AI API 公益中转站导航封面" width="1200" height="630">
    </div>
  </header>
  <section class="toolbar-band">
    <div class="toolbar wrap">
      <input id="search" class="search" type="search" placeholder="搜索站点、标签、模型或备注" autocomplete="off">
      <div class="filters" aria-label="标签筛选">
        <button class="filter active" type="button" data-filter="all">全部</button>
        {filters}
      </div>
    </div>
  </section>
  <main class="wrap">
    <div class="meta-row">
      <span id="count">显示 {len(sites)} 个站点</span>
      <span>请自行确认额度、倍率、模型和注册状态</span>
    </div>
    <section class="grid" id="sites" aria-label="AI API 公益中转站列表">
      {cards}
    </section>
    <p class="empty" id="empty">没有匹配的站点</p>
  </main>
  <footer class="wrap">
    <p>请保护个人隐私和数据安全；额度、倍率、模型和注册状态可能随时变化。本站只做 AI API 公益站和中转站导航收录，不对任何第三方站点服务质量负责。</p>
  </footer>
  <script>
    const searchInput = document.querySelector("#search");
    const cards = [...document.querySelectorAll(".site-card")];
    const count = document.querySelector("#count");
    const empty = document.querySelector("#empty");
    const filterButtons = [...document.querySelectorAll(".filter")];
    let activeFilter = "all";

    function applyFilters() {{
      const query = searchInput.value.trim().toLowerCase();
      let visible = 0;
      for (const card of cards) {{
        const matchQuery = !query || card.dataset.search.includes(query);
        const matchTag = activeFilter === "all" || card.dataset.tags.includes(activeFilter);
        const show = matchQuery && matchTag;
        card.hidden = !show;
        if (show) visible += 1;
      }}
      count.textContent = `显示 ${{visible}} 个站点`;
      empty.style.display = visible ? "none" : "block";
    }}

    searchInput.addEventListener("input", applyFilters);
    for (const button of filterButtons) {{
      button.addEventListener("click", () => {{
        activeFilter = button.dataset.filter;
        for (const item of filterButtons) item.classList.toggle("active", item === button);
        applyFilters();
      }});
    }}
  </script>
</body>
</html>
"""
    for key in ("index", "table_html", "share_html"):
        OUTPUTS[key].write_text(html_doc, encoding="utf-8")
    return html_doc


def write_search_files(updated: str) -> None:
    OUTPUTS["robots"].write_text(
        "\n".join(
            [
                "User-agent: *",
                "Allow: /",
                f"Sitemap: {SITE_URL}sitemap.xml",
                "",
            ]
        ),
        encoding="utf-8",
    )
    OUTPUTS["sitemap"].write_text(
        "\n".join(
            [
                '<?xml version="1.0" encoding="UTF-8"?>',
                '<urlset xmlns="http://www.sitemaps.org/schemas/sitemap/0.9">',
                "  <url>",
                f"    <loc>{esc(SITE_URL)}</loc>",
                f"    <lastmod>{esc(updated)}</lastmod>",
                "    <changefreq>weekly</changefreq>",
                "    <priority>1.0</priority>",
                "  </url>",
                "</urlset>",
                "",
            ]
        ),
        encoding="utf-8",
    )


def verify_links(html_doc: str, markdown_doc: str, sites: list[Site]) -> None:
    allowed = {site.url for site in sites}
    allowed_hrefs = allowed | {SITE_URL}
    hrefs = set(re.findall(r'href="([^"]+)"', html_doc))
    md_links = set(re.findall(r"\]\((https?://[^)]+)\)", markdown_doc))
    bad_hrefs = sorted(href for href in hrefs if href.startswith(("http://", "https://")) and href not in allowed_hrefs)
    bad_md = sorted(link for link in md_links if link not in allowed)
    if bad_hrefs or bad_md:
        raise ValueError(f"Unexpected non-aff links: html={bad_hrefs}, markdown={bad_md}")


def main() -> None:
    title, subtitle, sites = load_sites()
    updated = source_updated()
    write_cover(title, sites)
    write_csv(sites)
    write_json(title, subtitle, updated, sites)
    markdown_doc = write_markdown(title, subtitle, updated, sites)
    html_doc = write_html(title, subtitle, updated, sites)
    write_search_files(updated)
    verify_links(html_doc, markdown_doc, sites)
    print(f"Generated {len(sites)} sites from {SOURCE.name}")


if __name__ == "__main__":
    main()
