from __future__ import annotations

import argparse
import json
import re
import sys
from dataclasses import asdict, dataclass
from datetime import date
from pathlib import Path
from typing import Callable
from urllib.parse import urlparse

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
SOURCE_HEADERS = ("序号", "站点", "链接", "标签", "备注", "添加日期")
URL_PATTERN = re.compile(r"https?://[^\s<>\"']+", re.IGNORECASE)
REFERRAL_SIGNALS = ("aff=", "ref=", "invite=", "/invite/")
KNOWN_TAGS = (
    "公益",
    "签到",
    "生图",
    "稳定",
    "注册赠送",
    "低倍率",
    "Claude",
    "GPT",
    "DeepSeek",
    "Gemini",
    "GLM",
    "MiniMax",
    "Codex",
)
RELEASE_PATHS = (
    "ai-api-sites-table.xlsx",
    "index.html",
    "ai-api-sites-table.html",
    "ai-api-sites-share.html",
    "README.md",
    "ai-api-sites-table.md",
    "ai-api-sites-table.csv",
    "data/sites.json",
    "assets/ai-api-gongyi-nav-cover.png",
    "robots.txt",
    "sitemap.xml",
)


@dataclass(frozen=True)
class ParsedSite:
    name: str
    url: str
    tags: tuple[str, ...]
    note: str
    added_date: str


@dataclass(frozen=True)
class PublishResult:
    site: ParsedSite
    index: int
    dry_run: bool


def clean_url(value: str) -> str:
    return value.strip().rstrip(".,，。;；)）]】")


def field_values(copy_text: str) -> dict[str, str]:
    fields: dict[str, str] = {}
    aliases = {"站点": "name", "名称": "name", "链接": "url", "标签": "tags", "备注": "note"}
    for line in copy_text.splitlines():
        match = re.match(r"^\s*(站点|名称|链接|标签|备注)\s*[:：]\s*(.*?)\s*$", line)
        if match and match.group(2):
            fields[aliases[match.group(1)]] = match.group(2)
    return fields


def split_tags(value: str) -> tuple[str, ...]:
    result: list[str] = []
    for tag in re.split(r"[;；,，]", value):
        cleaned = tag.strip()
        if cleaned and cleaned not in result:
            result.append(cleaned)
    return tuple(result)


def derive_tags(copy_text: str) -> tuple[str, ...]:
    lowered = copy_text.lower()
    return tuple(tag for tag in KNOWN_TAGS if tag.lower() in lowered)


def first_url(copy_text: str) -> str:
    match = URL_PATTERN.search(copy_text)
    return clean_url(match.group(0)) if match else ""


def inferred_name(copy_text: str) -> str:
    for line in copy_text.splitlines():
        stripped = line.strip()
        if not stripped or URL_PATTERN.search(stripped):
            continue
        if re.match(r"^(站点|名称|链接|标签|备注)\s*[:：]", stripped):
            continue
        return stripped
    return ""


def inferred_note(copy_text: str, name: str) -> str:
    lines: list[str] = []
    for line in copy_text.splitlines():
        stripped = line.strip()
        if not stripped or stripped == name or URL_PATTERN.fullmatch(stripped):
            continue
        if re.match(r"^(站点|名称|链接|标签|备注)\s*[:：]", stripped):
            continue
        lines.append(stripped)
    return " ".join(lines)


def validate_url(url: str) -> None:
    parsed = urlparse(url)
    if parsed.scheme not in {"http", "https"} or not parsed.netloc:
        raise ValueError("URL must be a valid http or https URL")
    if not any(signal in url.lower() for signal in REFERRAL_SIGNALS):
        raise ValueError("referral signal is required in the URL")


def parse_site_copy(copy_text: str, added_date: str) -> ParsedSite:
    fields = field_values(copy_text)
    name = fields.get("name", inferred_name(copy_text)).strip()
    url = clean_url(fields.get("url", first_url(copy_text)))
    if not name:
        raise ValueError("name is required")
    if not url:
        raise ValueError("URL is required")
    validate_url(url)
    tags = split_tags(fields["tags"]) if "tags" in fields else derive_tags(copy_text)
    note = fields.get("note", inferred_note(copy_text, name)).strip()
    if not note:
        raise ValueError("note is required")
    return ParsedSite(name=name, url=url, tags=tags, note=note, added_date=added_date)


def normalized_domain(url: str) -> str:
    hostname = urlparse(url).hostname or ""
    return hostname.lower().removeprefix("www.")


def find_header_row(sheet: object) -> int:
    for row_number in range(1, min(sheet.max_row, 10) + 1):
        headers = tuple(sheet.cell(row=row_number, column=column).value for column in range(1, 7))
        if headers == SOURCE_HEADERS:
            return row_number
    raise ValueError("source workbook is missing the required headers")


def existing_rows(workbook_path: Path) -> tuple[int, list[tuple[int, str]]]:
    workbook = load_workbook(workbook_path)
    sheet = workbook.active
    header_row = find_header_row(sheet)
    rows: list[tuple[int, str]] = []
    for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
        if not any(row):
            continue
        index = int(row[0])
        url = str(row[2] or "")
        rows.append((index, url))
    return header_row, rows


def next_index(workbook_path: Path, site: ParsedSite) -> int:
    _, rows = existing_rows(workbook_path)
    domain = normalized_domain(site.url)
    if any(normalized_domain(url) == domain for _, url in rows):
        raise ValueError(f"domain already exists: {domain}")
    return max((index for index, _ in rows), default=0) + 1


def append_site(workbook_path: Path, site: ParsedSite) -> int:
    index = next_index(workbook_path, site)
    workbook = load_workbook(workbook_path)
    sheet = workbook.active
    sheet.append([index, site.name, site.url, ";".join(site.tags), site.note, site.added_date])
    workbook.save(workbook_path)
    return index


def publish(workbook_path: Path, site: ParsedSite, dry_run: bool, generate: Callable[[], None]) -> PublishResult:
    index = next_index(workbook_path, site)
    if dry_run:
        return PublishResult(site=site, index=index, dry_run=True)
    append_site(workbook_path, site)
    generate()
    return PublishResult(site=site, index=index, dry_run=False)


def result_payload(result: PublishResult) -> dict[str, object]:
    return {"site": asdict(result.site), "index": result.index, "dry_run": result.dry_run}


def main() -> int:
    parser = argparse.ArgumentParser(description="Publish one AI API site from promotional copy.")
    parser.add_argument("--input", type=Path, help="UTF-8 promotional copy file")
    parser.add_argument("--dry-run", action="store_true", help="Validate and preview without writing")
    parser.add_argument("--release-paths", action="store_true", help="Print paths that may be staged for release")
    args = parser.parse_args()

    if args.release_paths:
        print("\n".join(RELEASE_PATHS))
        return 0
    if args.input is None:
        parser.error("--input is required unless --release-paths is used")

    try:
        site = parse_site_copy(args.input.read_text(encoding="utf-8"), added_date=date.today().isoformat())
        from generate_site import main as generate_site

        result = publish(ROOT / "ai-api-sites-table.xlsx", site, args.dry_run, generate_site)
    except (OSError, ValueError) as error:
        print(str(error), file=sys.stderr)
        return 1
    print(json.dumps(result_payload(result), ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
