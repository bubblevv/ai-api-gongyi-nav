from __future__ import annotations

import sys
import tempfile
import unittest
from pathlib import Path
from subprocess import run

from openpyxl import Workbook, load_workbook

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "tools"))

import publish_site


class ParseCopyTests(unittest.TestCase):
    def test_parse_structured_copy_uses_explicit_fields(self) -> None:
        copy_text = """站点：星河 API
链接：https://api.example.com/register?aff=invite123
标签：公益;签到;GPT
备注：注册赠送 10 刀，0.1 倍率
"""

        site = publish_site.parse_site_copy(copy_text, added_date="2026-07-16")

        self.assertEqual(site.name, "星河 API")
        self.assertEqual(site.url, "https://api.example.com/register?aff=invite123")
        self.assertEqual(site.tags, ("公益", "签到", "GPT"))
        self.assertEqual(site.note, "注册赠送 10 刀，0.1 倍率")
        self.assertEqual(site.added_date, "2026-07-16")

    def test_parse_free_form_copy_derives_only_present_known_tags(self) -> None:
        copy_text = "星河 API\nhttps://api.example.com/register?ref=abc\n公益 GPT 注册赠送 10 刀"

        site = publish_site.parse_site_copy(copy_text, added_date="2026-07-16")

        self.assertEqual(site.name, "星河 API")
        self.assertEqual(site.tags, ("公益", "注册赠送", "GPT"))
        self.assertEqual(site.note, "公益 GPT 注册赠送 10 刀")

    def test_parse_rejects_url_without_referral_signal(self) -> None:
        with self.assertRaisesRegex(ValueError, "referral"):
            publish_site.parse_site_copy("站点：星河\nhttps://api.example.com/register", added_date="2026-07-16")

    def test_parse_rejects_missing_name(self) -> None:
        with self.assertRaisesRegex(ValueError, "name"):
            publish_site.parse_site_copy("https://api.example.com/register?aff=abc", added_date="2026-07-16")


class WorkbookTests(unittest.TestCase):
    def test_source_workbook_is_not_ignored(self) -> None:
        result = run(
            ["git", "check-ignore", "-q", "ai-api-sites-table.xlsx"],
            cwd=ROOT,
            check=False,
        )

        self.assertNotEqual(result.returncode, 0)

    def workbook(self) -> Path:
        directory = tempfile.TemporaryDirectory()
        self.addCleanup(directory.cleanup)
        path = Path(directory.name) / "sites.xlsx"
        book = Workbook()
        sheet = book.active
        sheet.append(["公益站导航"])
        sheet.append(["序号", "站点", "链接", "标签", "备注", "添加日期"])
        sheet.append([1, "已有站", "https://old.example/register?aff=old", "公益", "已有", "2026-07-01"])
        book.save(path)
        return path

    def test_append_site_adds_next_index_and_date(self) -> None:
        workbook = self.workbook()
        site = publish_site.parse_site_copy(
            "站点：新站\nhttps://new.example/register?aff=new\n标签：公益;GPT\n备注：注册送 1 刀",
            added_date="2026-07-16",
        )

        index = publish_site.append_site(workbook, site)
        row = list(load_workbook(workbook, data_only=True).active.iter_rows(values_only=True))[-1]

        self.assertEqual(index, 2)
        self.assertEqual(row, (2, "新站", site.url, "公益;GPT", "注册送 1 刀", "2026-07-16"))

    def test_append_site_rejects_existing_normalized_url(self) -> None:
        workbook = self.workbook()
        site = publish_site.parse_site_copy(
            "站点：重复站\nHTTPS://OLD.EXAMPLE/register/?aff=old\n备注：重复链接",
            added_date="2026-07-16",
        )

        with self.assertRaisesRegex(ValueError, "URL already exists"):
            publish_site.append_site(workbook, site)

    def test_append_site_allows_distinct_referral_url_on_existing_domain(self) -> None:
        workbook = self.workbook()
        site = publish_site.parse_site_copy(
            "站点：同域新链接\nhttps://old.example/register?aff=new\n备注：不同邀请链接",
            added_date="2026-07-16",
        )

        index = publish_site.append_site(workbook, site)

        self.assertEqual(index, 2)

    def test_dry_run_does_not_mutate_workbook(self) -> None:
        workbook = self.workbook()
        before = workbook.read_bytes()
        site = publish_site.parse_site_copy(
            "站点：预览站\nhttps://preview.example/register?aff=preview\n备注：仅预览",
            added_date="2026-07-16",
        )

        result = publish_site.publish(workbook, site, dry_run=True, generate=lambda: None)

        self.assertEqual(result.index, 2)
        self.assertEqual(before, workbook.read_bytes())

    def test_publish_restores_workbook_when_generator_fails(self) -> None:
        workbook = self.workbook()
        before = workbook.read_bytes()
        site = publish_site.parse_site_copy(
            "站点：失败站\nhttps://failure.example/register?aff=failure\n备注：生成失败",
            added_date="2026-07-16",
        )

        with self.assertRaisesRegex(RuntimeError, "generator failed"):
            publish_site.publish(
                workbook,
                site,
                dry_run=False,
                generate=lambda: (_ for _ in ()).throw(RuntimeError("generator failed")),
            )

        self.assertEqual(before, workbook.read_bytes())

    def test_publish_restores_generated_files_when_generator_fails(self) -> None:
        original_root = publish_site.ROOT
        with tempfile.TemporaryDirectory() as directory:
            output_root = Path(directory)
            generated_file = output_root / "index.html"
            generated_file.write_text("before", encoding="utf-8")
            publish_site.ROOT = output_root
            workbook = self.workbook()
            site = publish_site.parse_site_copy(
                "站点：失败站\nhttps://failure.example/register?aff=failure\n备注：生成失败",
                added_date="2026-07-16",
            )

            def failing_generator() -> None:
                generated_file.write_text("after", encoding="utf-8")
                raise RuntimeError("generator failed")

            try:
                with self.assertRaisesRegex(RuntimeError, "generator failed"):
                    publish_site.publish(workbook, site, dry_run=False, generate=failing_generator)
            finally:
                publish_site.ROOT = original_root

            self.assertEqual(generated_file.read_text(encoding="utf-8"), "before")

    def test_release_paths_are_explicit_and_exclude_reports(self) -> None:
        self.assertIn("ai-api-sites-table.xlsx", publish_site.RELEASE_PATHS)
        self.assertIn("data/sites.json", publish_site.RELEASE_PATHS)
        self.assertNotIn("reports", publish_site.RELEASE_PATHS)
        self.assertNotIn("tools/collect_candidates.py", publish_site.RELEASE_PATHS)


class SkillSafetyTests(unittest.TestCase):
    def test_skill_uses_publisher_git_preflight_before_publishing(self) -> None:
        skill = (ROOT / ".agents" / "skills" / "publish-ai-site" / "SKILL.md").read_text(encoding="utf-8")

        self.assertIn("python tools\\publish_site.py --preflight", skill)


class GitPreflightTests(unittest.TestCase):
    def repository(self) -> Path:
        directory = tempfile.TemporaryDirectory()
        self.addCleanup(directory.cleanup)
        repo = Path(directory.name)
        for command in (
            ["git", "init", "-q"],
            ["git", "config", "user.email", "test@example.com"],
            ["git", "config", "user.name", "Test User"],
        ):
            run(command, cwd=repo, check=True)
        return repo

    def test_git_preflight_rejects_staged_changes(self) -> None:
        repo = self.repository()
        (repo / "unrelated.txt").write_text("staged", encoding="utf-8")
        run(["git", "add", "unrelated.txt"], cwd=repo, check=True)

        with self.assertRaisesRegex(ValueError, "staged changes"):
            publish_site.require_clean_git_preflight(repo)

    def test_git_preflight_rejects_dirty_release_paths(self) -> None:
        repo = self.repository()
        (repo / "README.md").write_text("dirty", encoding="utf-8")

        with self.assertRaisesRegex(ValueError, "release paths already changed"):
            publish_site.require_clean_git_preflight(repo)

    def test_git_preflight_allows_unrelated_untracked_files(self) -> None:
        repo = self.repository()
        (repo / "collector.txt").write_text("untracked", encoding="utf-8")

        publish_site.require_clean_git_preflight(repo)


if __name__ == "__main__":
    unittest.main()
