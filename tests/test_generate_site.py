from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook

import sys

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "tools"))

import generate_site


class LoadSitesTests(unittest.TestCase):
    def setUp(self) -> None:
        self.original_source = generate_site.SOURCE
        self.original_outputs = generate_site.OUTPUTS.copy()

    def tearDown(self) -> None:
        generate_site.SOURCE = self.original_source
        generate_site.OUTPUTS.clear()
        generate_site.OUTPUTS.update(self.original_outputs)

    def write_workbook(self, headers: list[str], row: list[object]) -> Path:
        temp_dir = tempfile.TemporaryDirectory()
        self.addCleanup(temp_dir.cleanup)
        path = Path(temp_dir.name) / "sites.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet.cell(row=1, column=1).value = "公益站导航\n测试说明"
        for column, header in enumerate(headers, start=1):
            sheet.cell(row=2, column=column).value = header
        for column, value in enumerate(row, start=1):
            sheet.cell(row=3, column=column).value = value
        workbook.save(path)
        return path

    def test_load_sites_reads_added_date_column(self) -> None:
        generate_site.SOURCE = self.write_workbook(
            ["序号", "站点", "链接", "标签", "备注", "添加日期"],
            [1, "J3GB", "https://vip.j3gb.com/register?aff=KEHXUQAGYBF8", "公益;注册赠送", "注册送15刀", "2026-07-01"],
        )

        _, _, sites = generate_site.load_sites()

        self.assertEqual(sites[0].added_date, "2026-07-01")
        self.assertEqual(generate_site.site_to_dict(sites[0])["added_date"], "2026-07-01")

    def test_load_sites_keeps_old_five_column_workbooks_compatible(self) -> None:
        generate_site.SOURCE = self.write_workbook(
            ["序号", "站点", "链接", "标签", "备注"],
            [1, "Old", "https://old.example.com/register?aff=abc", "公益", "旧格式"],
        )

        _, _, sites = generate_site.load_sites()

        self.assertEqual(sites[0].added_date, "")

    def test_write_markdown_omits_generated_file_manifest(self) -> None:
        temp_dir = tempfile.TemporaryDirectory()
        self.addCleanup(temp_dir.cleanup)
        output_dir = Path(temp_dir.name)
        generate_site.OUTPUTS["readme"] = output_dir / "README.md"
        generate_site.OUTPUTS["markdown"] = output_dir / "sites.md"
        site = generate_site.Site(
            index=1,
            name="JianZhiLe",
            url="https://jianzhile.vip/register?aff=iszB",
            tags=("签到", "GPT"),
            note="【0706新增】注册20刀，签到20刀",
            added_date="2026-07-06",
        )

        content = generate_site.write_markdown("公益站导航", "测试说明", "2026-07-06", [site])

        self.assertIn("JianZhiLe", content)
        self.assertIn("添加日期", content)
        self.assertNotIn("## 文件", content)
        self.assertNotIn("`index.html`", content)
        self.assertNotIn("tools/generate_site.py", content)


if __name__ == "__main__":
    unittest.main()
